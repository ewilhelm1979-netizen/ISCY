import csv
import io
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Tuple
from urllib.request import Request, urlopen

from django.conf import settings
from openpyxl import load_workbook

from apps.assets_app.models import InformationAsset
from apps.organizations.models import BusinessUnit, Supplier
from apps.processes.models import Process


TRUE_VALUES = {"1", "true", "yes", "ja", "y"}

COLUMN_SYNONYMS = {
    'name': ['Name'],
    'owner_email': ['OwnerEmail', 'Verantwortlicher'],
    'scope': ['Scope'],
    'description': ['Beschreibung', 'Service'],
    'status': ['Status'],
    'business_unit': ['BusinessUnit', 'Geschäftsbereich'],
    'documented': ['Dokumentiert'],
    'implemented': ['Umgesetzt'],
    'evidenced': ['Nachweisbar'],
    'approved': ['Genehmigt'],
    'communicated': ['Kommuniziert'],
    'effective': ['Wirksam'],
    'service_description': ['Beschreibung', 'Service'],
    'criticality': ['Kritikalität'],
    'asset_type': ['Typ'],
    'confidentiality': ['Vertraulichkeit'],
    'integrity': ['Integrität'],
    'availability': ['Verfügbarkeit'],
    'lifecycle_status': ['Lifecycle'],
    'in_scope': ['ImScope'],
}

TEMPLATE_COLUMNS = {
    'business_units': ['name', 'owner_email'],
    'processes': ['name', 'scope', 'description', 'status', 'business_unit', 'documented', 'implemented', 'evidenced', 'approved', 'communicated', 'effective'],
    'suppliers': ['name', 'service_description', 'criticality'],
    'assets': ['name', 'asset_type', 'criticality', 'description', 'business_unit', 'confidentiality', 'integrity', 'availability', 'lifecycle_status', 'in_scope'],
}


@dataclass(frozen=True)
class ImportJobBridgeResult:
    tenant_id: int
    import_type: str
    row_count: int
    created: int
    updated: int
    skipped: int


class ImportRustClient:
    @staticmethod
    def _base_url() -> str:
        base = (getattr(settings, 'RUST_BACKEND_URL', '') or '').strip()
        return base.rstrip('/')

    @staticmethod
    def apply_import(
        request,
        tenant,
        import_type: str,
        rows: list[dict],
        replace_existing: bool = False,
        timeout: int = 30,
    ) -> ImportJobBridgeResult:
        base = ImportRustClient._base_url()
        if not base:
            raise RuntimeError('RUST_BACKEND_URL ist nicht gesetzt.')

        payload = {
            'import_type': import_type,
            'replace_existing': bool(replace_existing),
            'rows': rows,
        }
        rust_request = ImportRustClient._authenticated_request(
            request,
            tenant,
            f'{base}/api/v1/import-center/jobs',
            payload,
        )
        with urlopen(rust_request, timeout=timeout) as response:
            response_payload = json.loads(response.read().decode('utf-8'))

        result = response_payload.get('result') or {}
        tenant_id = int(result.get('tenant_id') or 0)
        if tenant_id != int(tenant.id):
            raise RuntimeError('Rust-Importjob gehoert nicht zum angefragten Tenant.')

        return ImportJobBridgeResult(
            tenant_id=tenant_id,
            import_type=str(result.get('import_type') or import_type),
            row_count=int(result.get('row_count') or 0),
            created=int(result.get('created') or 0),
            updated=int(result.get('updated') or 0),
            skipped=int(result.get('skipped') or 0),
        )

    @staticmethod
    def _authenticated_request(request, tenant, url: str, payload: dict) -> Request:
        user = getattr(request, 'user', None)
        user_id = getattr(user, 'id', None)
        if not user_id:
            raise RuntimeError('Import-Rust-Bridge braucht einen authentifizierten User.')

        body = json.dumps(payload).encode('utf-8')
        rust_request = Request(url, data=body, method='POST')
        rust_request.add_header('Accept', 'application/json')
        rust_request.add_header('Content-Type', 'application/json')
        rust_request.add_header('X-ISCY-Tenant-ID', str(tenant.id))
        rust_request.add_header('X-ISCY-User-ID', str(user_id))
        user_email = (getattr(user, 'email', '') or '').strip()
        if user_email:
            rust_request.add_header('X-ISCY-User-Email', user_email)
        return rust_request


class ImportCenterBridge:
    @staticmethod
    def apply_import(request, tenant, import_type: str, rows: list[dict], replace_existing: bool = False):
        if tenant is None:
            return None

        backend = str(getattr(settings, 'IMPORT_CENTER_BACKEND', 'rust_service') or '').strip().lower()
        if backend != 'rust_service':
            return None

        if not ImportRustClient._base_url():
            if ImportCenterBridge._strict_rust_mode():
                raise RuntimeError('Rust import backend ist aktiv, aber RUST_BACKEND_URL ist nicht gesetzt.')
            return None

        try:
            return ImportRustClient.apply_import(
                request,
                tenant,
                import_type,
                rows,
                replace_existing=replace_existing,
            )
        except Exception as exc:
            if ImportCenterBridge._strict_rust_mode():
                raise RuntimeError('Rust import backend ist aktiv, aber nicht erreichbar.') from exc
            return None

    @staticmethod
    def _strict_rust_mode() -> bool:
        return bool(getattr(settings, 'RUST_STRICT_MODE', False))


class ImportService:
    @staticmethod
    def read_rows(uploaded_file) -> Tuple[List[str], List[dict]]:
        suffix = Path(uploaded_file.name).suffix.lower()
        if suffix == '.csv':
            data = uploaded_file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(data))
            rows = [row for row in reader]
            return reader.fieldnames or [], rows
        if suffix in {'.xlsx', '.xlsm'}:
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
            values = list(ws.iter_rows(values_only=True))
            if not values:
                return [], []
            headers = [str(v).strip() if v is not None else '' for v in values[0]]
            rows = []
            for row in values[1:]:
                item = {}
                for idx, header in enumerate(headers):
                    item[header] = row[idx] if idx < len(row) else None
                rows.append(item)
            return headers, rows
        raise ValueError('Nur CSV oder XLSX werden unterstützt.')

    @staticmethod
    def bool_value(value) -> bool:
        return str(value).strip().lower() in TRUE_VALUES

    @staticmethod
    def normalize(value):
        return str(value).strip() if value is not None else ''

    @staticmethod
    def expected_columns(import_type: str):
        return TEMPLATE_COLUMNS.get(import_type, [])

    @staticmethod
    def default_mapping(import_type: str, headers: list[str]):
        expected = ImportService.expected_columns(import_type)
        normalized = {str(h).strip().lower(): h for h in headers}
        mapping = {}
        for canonical in expected:
            matched = normalized.get(canonical.lower())
            if not matched:
                for synonym in COLUMN_SYNONYMS.get(canonical, []):
                    if synonym.lower() in normalized:
                        matched = normalized[synonym.lower()]
                        break
            mapping[canonical] = matched or ''
        return mapping

    @staticmethod
    def get_mapping_preview(import_type: str, headers: list[str], selected_mapping: dict | None = None):
        expected = TEMPLATE_COLUMNS.get(import_type, [])
        selected_mapping = selected_mapping or ImportService.default_mapping(import_type, headers)
        rows = []
        selected_values = {str(v).strip().lower() for v in selected_mapping.values() if v}
        extra_headers = [h for h in headers if str(h).strip().lower() not in selected_values]
        for canonical in expected:
            matched = selected_mapping.get(canonical, '')
            rows.append({
                'expected': canonical,
                'matched': matched,
                'status': 'ok' if matched else 'missing',
                'synonyms': COLUMN_SYNONYMS.get(canonical, []),
                'required': canonical == 'name',
            })
        return rows, extra_headers

    @staticmethod
    def apply_mapping(rows: list[dict], selected_mapping: dict):
        transformed = []
        for row in rows:
            out = {}
            for expected, source in selected_mapping.items():
                out[expected] = row.get(source) if source else None
            transformed.append(out)
        return transformed

    @staticmethod
    def import_business_units(tenant, rows: Iterable[dict], replace_existing=False):
        if replace_existing:
            BusinessUnit.objects.filter(tenant=tenant).delete()
        created = updated = 0
        for row in rows:
            name = ImportService.normalize(row.get('name') or row.get('Name'))
            if not name:
                continue
            _, was_created = BusinessUnit.objects.update_or_create(
                tenant=tenant,
                name=name,
                defaults={},
            )
            created += int(was_created)
            updated += int(not was_created)
        return created, updated

    @staticmethod
    def import_processes(tenant, rows: Iterable[dict], replace_existing=False):
        if replace_existing:
            Process.objects.filter(tenant=tenant).delete()
        created = updated = 0
        status_map = {choice[0].lower(): choice[0] for choice in Process.Status.choices}
        for row in rows:
            name = ImportService.normalize(row.get('name') or row.get('Name'))
            if not name:
                continue
            scope = ImportService.normalize(row.get('scope') or row.get('Scope'))
            description = ImportService.normalize(row.get('description') or row.get('Beschreibung'))
            status_raw = ImportService.normalize(row.get('status') or row.get('Status')).lower()
            status = status_map.get(status_raw, Process.Status.MISSING)
            business_unit_name = ImportService.normalize(row.get('business_unit') or row.get('BusinessUnit') or row.get('Geschäftsbereich'))
            business_unit = None
            if business_unit_name:
                business_unit, _ = BusinessUnit.objects.get_or_create(tenant=tenant, name=business_unit_name)
            obj, was_created = Process.objects.update_or_create(
                tenant=tenant,
                name=name,
                defaults={
                    'scope': scope,
                    'description': description,
                    'status': status,
                    'business_unit': business_unit,
                    'documented': ImportService.bool_value(row.get('documented') or row.get('Dokumentiert')),
                    'implemented': ImportService.bool_value(row.get('implemented') or row.get('Umgesetzt')),
                    'evidenced': ImportService.bool_value(row.get('evidenced') or row.get('Nachweisbar')),
                    'approved': ImportService.bool_value(row.get('approved') or row.get('Genehmigt')),
                    'communicated': ImportService.bool_value(row.get('communicated') or row.get('Kommuniziert')),
                    'effective': ImportService.bool_value(row.get('effective') or row.get('Wirksam')),
                },
            )
            created += int(was_created)
            updated += int(not was_created)
        return created, updated

    @staticmethod
    def import_suppliers(tenant, rows: Iterable[dict], replace_existing=False):
        if replace_existing:
            Supplier.objects.filter(tenant=tenant).delete()
        created = updated = 0
        for row in rows:
            name = ImportService.normalize(row.get('name') or row.get('Name'))
            if not name:
                continue
            obj, was_created = Supplier.objects.update_or_create(
                tenant=tenant,
                name=name,
                defaults={
                    'service_description': ImportService.normalize(row.get('service_description') or row.get('Beschreibung') or row.get('Service')),
                    'criticality': (ImportService.normalize(row.get('criticality') or row.get('Kritikalität')) or 'MEDIUM').upper(),
                },
            )
            created += int(was_created)
            updated += int(not was_created)
        return created, updated

    @staticmethod
    def import_assets(tenant, rows: Iterable[dict], replace_existing=False):
        if replace_existing:
            InformationAsset.objects.filter(tenant=tenant).delete()
        created = updated = 0
        type_map = {choice[0].lower(): choice[0] for choice in InformationAsset.Type.choices}
        crit_map = {choice[0].lower(): choice[0] for choice in InformationAsset.Criticality.choices}
        for row in rows:
            name = ImportService.normalize(row.get('name') or row.get('Name'))
            if not name:
                continue
            bu_name = ImportService.normalize(row.get('business_unit') or row.get('Geschäftsbereich'))
            business_unit = None
            if bu_name:
                business_unit, _ = BusinessUnit.objects.get_or_create(tenant=tenant, name=bu_name)
            obj, was_created = InformationAsset.objects.update_or_create(
                tenant=tenant,
                name=name,
                defaults={
                    'business_unit': business_unit,
                    'asset_type': type_map.get(ImportService.normalize(row.get('asset_type') or row.get('Typ')).lower(), InformationAsset.Type.APPLICATION),
                    'criticality': crit_map.get(ImportService.normalize(row.get('criticality') or row.get('Kritikalität')).lower(), InformationAsset.Criticality.MEDIUM),
                    'description': ImportService.normalize(row.get('description') or row.get('Beschreibung')),
                    'confidentiality': ImportService.normalize(row.get('confidentiality') or row.get('Vertraulichkeit')),
                    'integrity': ImportService.normalize(row.get('integrity') or row.get('Integrität')),
                    'availability': ImportService.normalize(row.get('availability') or row.get('Verfügbarkeit')),
                    'lifecycle_status': ImportService.normalize(row.get('lifecycle_status') or row.get('Lifecycle')),
                    'is_in_scope': not (ImportService.normalize(row.get('in_scope') or row.get('ImScope')).lower() in {'0', 'false', 'no', 'nein'}),
                },
            )
            created += int(was_created)
            updated += int(not was_created)
        return created, updated
